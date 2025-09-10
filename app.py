import streamlit as st
import pandas as pd
import difflib
import time
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Comparador de Planilhas", page_icon="📊")
st.title("Comparador de Planilhas")
st.markdown("Preencha o formulário inicial para configurar a comparação.")

# === FORMULÁRIO INICIAL ===
with st.form("config_form"):
    n_planilhas = st.number_input("Quantas planilhas você deseja comparar?", min_value=1, max_value=5, value=1, step=1)
    n_colunas = st.number_input("Quantas colunas devem ser comparadas em cada planilha?", min_value=1, max_value=5, value=1, step=1)
    modo = None
    if n_planilhas > 1:
        modo = st.radio("Modo de comparação", ["Primeira planilha contra as outras", "Comparação cruzada entre todas"])
    iniciar = st.form_submit_button("Continuar")

if iniciar:
    st.session_state["n_planilhas"] = n_planilhas
    st.session_state["n_colunas"] = n_colunas
    st.session_state["modo"] = modo

if "n_planilhas" in st.session_state:
    n_planilhas = st.session_state["n_planilhas"]
    n_colunas = st.session_state["n_colunas"]
    modo = st.session_state.get("modo")

    st.subheader("Upload das planilhas")
    arquivos = []
    for i in range(n_planilhas):
        uploaded = st.file_uploader(f"Selecione a planilha {i+1}", type=["xlsx"], key=f"file_{i}")
        arquivos.append(uploaded)

    if all(arquivos):
        planilhas = [pd.ExcelFile(f) for f in arquivos]

        abas = []
        dfs = []
        for i, p in enumerate(planilhas):
            aba = st.selectbox(f"Escolha a aba da planilha {i+1}", p.sheet_names, key=f"aba_{i}")
            abas.append(aba)
            dfs.append(p.parse(aba))

        st.subheader("Configuração das colunas")
        colunas_escolhidas = []
        for i, df in enumerate(dfs):
            cols = []
            for j in range(n_colunas):
                col = st.selectbox(f"Coluna {j+1} da planilha {i+1}", df.columns, key=f"col_{i}_{j}")
                cols.append(col)
            colunas_escolhidas.append(cols)

        def encontrar_mais_similar(texto, lista_textos):
            texto = str(texto)
            lista_textos = [str(t) for t in lista_textos]
            similar = difflib.get_close_matches(texto, lista_textos, n=1, cutoff=0.0)
            if not similar:
                return "Nenhum encontrado", "N/A"
            mais_similar = similar[0]

            d = difflib.Differ()
            diff = list(d.compare(texto.split(), mais_similar.split()))
            diffs = [x for x in diff if x.startswith("+ ") or x.startswith("- ")]
            diff_text = "Nenhuma diferença" if not diffs else " | ".join(diffs)
            return mais_similar, diff_text

        if st.button("Iniciar Comparação"):
            resultados_finais = []
            total_passos = 0

            if n_planilhas == 1:
                total_passos = len(dfs[0]) * n_colunas
            elif modo == "Primeira planilha contra as outras":
                total_passos = len(dfs[0]) * n_colunas * (n_planilhas - 1)
            else:
                for i in range(n_planilhas):
                    for j in range(i + 1, n_planilhas):
                        total_passos += len(dfs[i]) * n_colunas

            progresso_bar = st.progress(0)
            progresso_text = st.empty()
            progresso_atual = 0
            inicio = time.time()

            if n_planilhas == 1:
                base_df = dfs[0].copy()
                for c in colunas_escolhidas[0]:
                    similares, diffs = [], []
                    valores = base_df[c].tolist()
                    for i, val in enumerate(valores):
                        similar, diff = encontrar_mais_similar(val, valores)
                        similares.append(similar)
                        diffs.append(diff)
                        progresso_atual += 1
                        progresso_bar.progress(progresso_atual / total_passos)
                        tempo_passado = time.time() - inicio
                        tempo_estimado = (tempo_passado / progresso_atual) * (total_passos - progresso_atual)
                        tempo_str = f"{tempo_estimado/60:.1f} minutos restantes" if tempo_estimado > 60 else f"{tempo_estimado:.1f} segundos restantes"
                        progresso_text.markdown(f"**Processando {progresso_atual}/{total_passos} itens... | {tempo_str}**")
                    base_df[f"Similar_{c}"] = similares
                    base_df[f"Diferenças_{c}"] = diffs
                resultados_finais.append(("Planilha Única", base_df))

            elif modo == "Primeira planilha contra as outras":
                base_df = dfs[0].copy()
                for idx, df in enumerate(dfs[1:], start=2):
                    for c in colunas_escolhidas[0]:
                        similares, diffs = [], []
                        for i, val in enumerate(base_df[c].tolist()):
                            similar, diff = encontrar_mais_similar(val, df[colunas_escolhidas[idx-1][0]])
                            similares.append(similar)
                            diffs.append(diff)
                            progresso_atual += 1
                            progresso_bar.progress(progresso_atual / total_passos)
                            tempo_passado = time.time() - inicio
                            tempo_estimado = (tempo_passado / progresso_atual) * (total_passos - progresso_atual)
                            tempo_str = f"{tempo_estimado/60:.1f} minutos restantes" if tempo_estimado > 60 else f"{tempo_estimado:.1f} segundos restantes"
                            progresso_text.markdown(f"**Processando {progresso_atual}/{total_passos} itens... | {tempo_str}**")
                        base_df[f"Similar_Planilha{idx}_{c}"] = similares
                        base_df[f"Diferenças_Planilha{idx}_{c}"] = diffs
                resultados_finais.append(("Planilha 1 vs Outras", base_df))

            else:
                for i, dfA in enumerate(dfs):
                    for j, dfB in enumerate(dfs):
                        if i < j:
                            temp_df = dfA.copy()
                            for c in colunas_escolhidas[i]:
                                similares, diffs = [], []
                                for val in temp_df[c].tolist():
                                    similar, diff = encontrar_mais_similar(val, dfB[colunas_escolhidas[j][0]])
                                    similares.append(similar)
                                    diffs.append(diff)
                                    progresso_atual += 1
                                    progresso_bar.progress(progresso_atual / total_passos)
                                    tempo_passado = time.time() - inicio
                                    tempo_estimado = (tempo_passado / progresso_atual) * (total_passos - progresso_atual)
                                    tempo_str = f"{tempo_estimado/60:.1f} minutos restantes" if tempo_estimado > 60 else f"{tempo_estimado:.1f} segundos restantes"
                                    progresso_text.markdown(f"**Processando {progresso_atual}/{total_passos} itens... | {tempo_str}**")
                                temp_df[f"Similar_Planilha{j+1}_{c}"] = similares
                                temp_df[f"Diferenças_Planilha{j+1}_{c}"] = diffs
                            resultados_finais.append((f"Planilha {i+1} vs {j+1}", temp_df))

            # === Exportar com formatação ===
            output = BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                for nome, df in resultados_finais:
                    df.to_excel(writer, sheet_name=nome[:31], index=False)
            output.seek(0)

            wb = load_workbook(output)
            for ws in wb.worksheets:
                for col in ws.columns:
                    col_letter = get_column_letter(col[0].column)
                    ws.column_dimensions[col_letter].width = 50
                    for i, cell in enumerate(col):
                        if i == 0:
                            cell.fill = PatternFill(start_color="ededed", end_color="ededed", fill_type="solid")
                            cell.font = Font(bold=True)
                        else:
                            cell.alignment = Alignment(wrap_text=True, vertical='top')

            output_final = BytesIO()
            wb.save(output_final)
            output_final.seek(0)

            st.success("✅ Comparação concluída com sucesso!")
            st.download_button(
                label="📥 Baixar planilha com diferenças",
                data=output_final,
                file_name="Diferenças.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
