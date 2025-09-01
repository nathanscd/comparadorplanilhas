import pandas as pd
import os
import difflib
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

df = pd.read_excel("seuarquivo.xlsx")

#Comparação
def comparar_resumido(Requisito, Similar_Text):
    if pd.isna(Coluna01) or pd.isna(Coluna02):
        return "Um dos textos está vazio"

    d = difflib.Differ()
    diff = list(d.compare(str(Coluna01).split(), str(Coluna02).split()))
    
    removidos = sorted(set(x[2:] for x in diff if x.startswith("- ")))
    adicionados = sorted(set(x[2:] for x in diff if x.startswith("+ ")))

    if not removidos and not adicionados:
        return "Nenhuma diferença encontrada"

    resumo = ""
    if removidos:
        resumo += "Removido: " + ", ".join(removidos[:20])
        if len(removidos) > 20:
            resumo += ", ..."
        resumo += " | "
    if adicionados:
        resumo += "Adicionado: " + ", ".join(adicionados[:20])
        if len(adicionados) > 20:
            resumo += ", ..."

    return resumo

df["Coluna_comparação"] = df.apply(lambda row: comparar_resumido(row["Coluna01"], row["Coluna02"]), axis=1)

arquivo_saida = "src.xlsx"
if os.path.exists(arquivo_saida):
    print(f"O arquivo {arquivo_saida} já existe. Ele será sobrescrito.")

df.to_excel(arquivo_saida, index=False)
print(f"Comparação concluída! Arquivo gerado: {arquivo_saida}")

wb = load_workbook(arquivo_saida)

#Formatação do arquivo
ws = wb.active

for col in ws.columns:
    col_letter = get_column_letter(col[0].column)
    ws.column_dimensions[col_letter].width = 50 
    for i, cell in enumerate(col):
        if i == 0:
            cell.fill = PatternFill(start_color="ededed", end_color="ededed", fill_type="solid")
            cell.font = Font(bold=True)
        else:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

wb.save("Diferenças.xlsx")
print("Arquivo formatado e salvo como Diferenças.xlsx")

